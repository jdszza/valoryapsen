"""
APSEN - Computador Central v3.2
Orquestrador ativo: recebe OS do order-generator, comanda adapters,
consolida eventos, persiste no DB, serve dashboard/IHM via REST + WebSocket.

Comunicação: REST/HTTP/WebSocket — sem MQTT.

Novidades v3.2:
  - Weight adapter + HX711 simulator (balança de mesa)
  - Triple Check: valida dispenser × câmera_mesa × balança após cada dispensa
    → 2+ divergências ativam trava de emergência (bloqueia OS)
  - GET /api/v1/trava — estado da trava
  - POST /api/v1/admin/liberar-trava — libera trava (role admin)
  - GET /api/v1/visao/historico — histórico de leituras CV
  - POST /api/v1/eventos/peso — eventos da balança HX711

Endpoints de entrada (dos adapters e order-generator):
  POST /api/v1/ordens              ← order-generator
  POST /api/v1/eventos/dispenser   ← dispenser-adapter
  POST /api/v1/eventos/cnc         ← cnc-adapter
  POST /api/v1/eventos/visao       ← vision-adapter
  POST /api/v1/eventos/peso        ← weight-adapter

Endpoints de leitura (dashboard, ihm_web):
  GET  /estado, /os/*, /dispensers/estado, /medicamentos, ...
  GET  /api/v1/trava
  GET  /api/v1/visao/historico
  WS   /ws
"""
import asyncio
import collections
import copy
import csv
import io
import itertools
import json
import logging
import threading
import time
from contextlib import asynccontextmanager
from datetime import datetime, timezone
from typing import Optional

from fastapi import Depends, FastAPI, HTTPException, status, WebSocket, WebSocketDisconnect
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, StreamingResponse
from fastapi.security import HTTPAuthorizationCredentials, HTTPBearer
from pydantic import BaseModel

import orchestrator as orch
from auth import criar_token, decodificar_token, verificar_senha
from config import settings, validar_secret_key
from database import (
    atualizar_item_os, atualizar_status_ordem,
    get_historico_visao, salvar_leitura_visao,
    atualizar_usuario, criar_usuario, expurgar_dados_antigos,
    get_alarmes, get_alarmes_por_os, get_cnc_recentes, get_dispensas, get_dispensas_recentes,
    get_dispensers_estado,
    get_historico_ordens, get_historico_sensor, get_log_manutencao,
    get_ordem_ativa, get_ordem_por_id, get_total_alarmes_ativos, get_ultimas_leituras,
    get_usuario, get_usuarios,
    init_db, limpar_dispenser_estado, listar_categorias, listar_medicamentos,
    resolver_alarme,
    salvar_alarme, salvar_cnc_evento, salvar_dispensa, salvar_dispenser_estado,
    salvar_leitura_sensor, salvar_manutencao, salvar_ordem,
    toggle_usuario_ativo,
)

logging.basicConfig(level=logging.INFO,
                    format="%(asctime)s [CENTRAL] %(levelname)s %(message)s")
logger = logging.getLogger(__name__)

# ── Estado em memória ──────────────────────────────────────────────────────────
_estado = {
    "cnc": {
        "status":         "idle",
        "os_id":          None,
        "dispenser_alvo": None,
        "posicao_x":      0.0,
        "posicao_y":      0.0,
        "ciclo_atual":    0,
        "total_ciclos":   0,
    },
    "os_ativa":     None,
    "dispensers": {
        str(i): {
            "status":                "idle",
            "medicamento":           None,
            "sku":                   None,
            "categoria":             None,
            "quantidade":            0,
            "quantidade_alvo":       0,
            "quantidade_dispensada": 0,
            "quantidade_residual":   0,
            "os_id":                 None,
        }
        for i in range(1, 7)
    },
    "fila_os":       [],
    "fila_tamanho":  0,
    # Teto de OS esperando (MAX_FILA_OS). Vai no payload para o dashboard poder
    # mostrar "3/5" em vez de um número sem escala.
    "fila_capacidade": settings.MAX_FILA_OS,
    "atribuicao_ia": [],
    # Derivado do banco por _contar_alarmes_ativos() — nunca incrementado à mão.
    "alarmes_ativos": 0,
    # Trava de Triple Check (bloqueio de OS até intervenção de supervisor)
    "trava": {
        "ativa":   False,
        "os_id":   None,
        "slot_id": None,
        "motivo":  "",
    },
    # Última leitura da balança HX711
    "peso": {
        "ultima_leitura":  None,
        "slot_id":         None,
        "ts":              None,
    },
    # Última leitura de cada câmera (atualizado por _handle_evento_visao)
    "visao": {
        "camera_dispenser": {
            "ultima_leitura": None,   # tipo do último evento
            "slot_id":        None,
            "match_sku":      None,
            "confianca":      None,
            "ts":             None,
        },
        "camera_mesa": {
            "ultima_leitura":      None,
            "slot_id":             None,
            "quantidade_detectada": None,
            "quantidade_esperada":  None,
            "confianca":           None,
            "ts":                  None,
        },
    },
}
_lock = threading.Lock()
_log_eventos: collections.deque = collections.deque(maxlen=100)

# ── Event loop global (set na lifespan) ───────────────────────────────────────
_loop: Optional[asyncio.AbstractEventLoop] = None


def _ts() -> str:
    return datetime.now(timezone.utc).isoformat()


def _log(tipo: str, msg: str, dados: dict = None):
    _log_eventos.appendleft({"tipo": tipo, "msg": msg, "dados": dados or {}, "ts": _ts()})


# ── WebSocket Manager ──────────────────────────────────────────────────────────
class WSManager:
    def __init__(self):
        self.active: list[WebSocket] = []

    async def connect(self, ws: WebSocket):
        await ws.accept()
        self.active.append(ws)

    def disconnect(self, ws: WebSocket):
        if ws in self.active:
            self.active.remove(ws)

    async def broadcast(self, data: dict):
        msg = json.dumps(data, default=str)
        dead = []
        for ws in self.active:
            try:
                await ws.send_text(msg)
            except Exception:
                dead.append(ws)
        for ws in dead:
            self.disconnect(ws)


_ws_manager = WSManager()
_broadcast_q: asyncio.Queue = asyncio.Queue()


async def _broadcast_worker():
    while True:
        data = await _broadcast_q.get()
        await _ws_manager.broadcast(data)


def _enfileirar_broadcast(data: dict):
    """Thread-safe: pode ser chamado de qualquer thread."""
    if _loop and not _loop.is_closed():
        _loop.call_soon_threadsafe(_broadcast_q.put_nowait, data)


# ── Throttle do broadcast ─────────────────────────────────────────────────────
# Cada broadcast custa um `copy.deepcopy` do estado inteiro + serialização JSON
# + envio para todo cliente conectado. O `movendo` da CNC chega a cada 0.5s
# durante todo o movimento, e a telemetria são 12 eventos a cada 15s só do
# dispenser: pagar o preço completo por evento desses é desperdício puro, já
# que o conteúdo muda pouco e ninguém perde informação vendo 2 quadros por
# segundo em vez de 30.
#
# Eventos de alta frequência passam por `prioritario=False`: se o último envio
# foi há menos de BROADCAST_MIN_INTERVALO_MS, só marcam pendência e o
# `_broadcast_flusher` manda o snapshot corrente no próximo tique. Transição de
# verdade — trava, fim de OS, alarme, carga, dispensa — sai na hora, sem
# throttle, porque atraso ali é atraso de decisão do operador.
_BROADCAST_MIN_INTERVALO_S = settings.BROADCAST_MIN_INTERVALO_MS / 1000.0
_throttle_lock = threading.Lock()
_ultimo_broadcast: float = 0.0
_broadcast_pendente: bool = False

# Tipos periódicos, emitidos pelo equipamento independentemente de qualquer OS.
_TIPOS_ALTA_FREQUENCIA = frozenset({"movendo", "telemetria", "status"})

# Amostragem de trajetória: 1 em N "movendo" vira linha em `cnc_eventos`.
_contador_movendo = itertools.count(1)


def _amostrar_movendo() -> bool:
    """True quando este "movendo" é o escolhido da amostra (0 = nunca)."""
    n = settings.CNC_AMOSTRAGEM_MOVENDO
    return n > 0 and next(_contador_movendo) % n == 0


def _broadcast_estado(prioritario: bool = True):
    """Publica o estado. `prioritario=False` respeita o throttle."""
    global _ultimo_broadcast, _broadcast_pendente

    with _throttle_lock:
        agora = time.monotonic()
        if not prioritario and (agora - _ultimo_broadcast) < _BROADCAST_MIN_INTERVALO_S:
            _broadcast_pendente = True
            return
        _ultimo_broadcast = agora
        _broadcast_pendente = False

    with _lock:
        snap = copy.deepcopy(_estado)
    _enfileirar_broadcast({"tipo": "estado", **snap})


async def _broadcast_flusher():
    """Envia o snapshot que o throttle segurou — no máximo um por intervalo.

    Sem isso, a última posição antes de uma pausa ficaria retida até o próximo
    evento: o dashboard mostraria a CNC parada onde ela não está mais.
    """
    while True:
        await asyncio.sleep(_BROADCAST_MIN_INTERVALO_S)
        with _throttle_lock:
            pendente = _broadcast_pendente
        if pendente:
            _broadcast_estado()


# ── Helpers DB async ──────────────────────────────────────────────────────────
async def _db(fn, *args):
    """Executa função síncrona de DB em threadpool — não bloqueia o event loop."""
    try:
        await asyncio.to_thread(fn, *args)
    except Exception as exc:
        logger.warning("[DB] %s(%s): %s", fn.__name__, args[:2], exc)


# ── Alarmes ativos: derivado do banco, com cache curto ────────────────────────
# `alarmes_ativos` já foi um contador em memória incrementado nos handlers. Ele
# só subia: resolver um alarme pela IHM não o baixava, e um restart o zerava
# mesmo com alarmes abertos no banco. Agora o valor é sempre uma leitura de
# `get_total_alarmes_ativos()`.
#
# O número vai em TODO payload de `_broadcast_estado()`, que roda a cada evento
# — só a telemetria periódica são 6 slots a cada 15s, mais CNC, visão e peso.
# Uma query por evento seria desperdício, então a leitura é cacheada por
# `_ALARMES_TTL_S`.
#
# Por que TTL, e não invalidar só quando o central cria/resolve um alarme: o
# central NÃO é o único a escrever na tabela. `orchestrator.py` grava alarmes
# de trava do Triple Check, de abort de OS e de falha de limpeza chamando
# `salvar_alarme` direto, sem passar por aqui — um cache invalidado apenas
# pelos caminhos deste módulo ficaria permanentemente atrasado em relação a
# eles. O TTL converge sozinho, seja quem for que escreveu. Os pontos que o
# central conhece (handler que abriu alarme, resolução pela IHM, startup)
# passam `forcar=True` para não fazer o badge esperar a janela.
_ALARMES_TTL_S = 5.0
_alarmes_lock = threading.Lock()
_alarmes_cache = {"valor": 0, "lido_em": None}   # lido_em None = nunca lido


def _alarmes_em_cache() -> Optional[int]:
    """Valor cacheado ainda válido, ou None se nunca lido / expirado."""
    with _alarmes_lock:
        lido_em = _alarmes_cache["lido_em"]
        if lido_em is None or (time.monotonic() - lido_em) >= _ALARMES_TTL_S:
            return None
        return _alarmes_cache["valor"]


def _contar_alarmes_ativos(forcar: bool = False) -> int:
    """Relê o total de alarmes abertos e publica em `_estado`.

    Bloqueante (toca o banco): chamar de endpoint síncrono ou via
    `asyncio.to_thread`. Nunca chamar com `_lock` em mãos — a função o adquire.
    """
    if not forcar:
        cacheado = _alarmes_em_cache()
        if cacheado is not None:
            return cacheado

    try:
        total = int(get_total_alarmes_ativos() or 0)
    except Exception as exc:
        # Banco fora do ar não pode derrubar o handler do evento: mantém o
        # último valor conhecido e tenta de novo na próxima chamada.
        logger.warning("[DB] get_total_alarmes_ativos: %s", exc)
        with _alarmes_lock:
            return _alarmes_cache["valor"]

    with _alarmes_lock:
        _alarmes_cache.update({"valor": total, "lido_em": time.monotonic()})
    with _lock:
        _estado["alarmes_ativos"] = total
    return total


async def _atualizar_alarmes_ativos(forcar: bool = False) -> int:
    """Versão para o event loop: só vai ao threadpool se o cache não servir."""
    if not forcar:
        cacheado = _alarmes_em_cache()
        if cacheado is not None:
            return cacheado
    return await asyncio.to_thread(_contar_alarmes_ativos, forcar)


def _abriu_alarme(db_tasks: list) -> bool:
    """Algum dos writes deste evento é um alarme novo?

    Lido das próprias `db_tasks` em vez de uma flag por handler: ponto de
    alarme novo entra na conta sozinho, sem depender de alguém lembrar.
    """
    return any(fn is salvar_alarme for fn, _ in db_tasks)


# ── Handlers de eventos vindos dos adapters ────────────────────────────────────

async def _handle_evento_dispenser(payload: dict):
    """
    Processa eventos do dispenser-adapter e notifica o orquestrador.
    Atualização de estado (rápida, sob _lock) é síncrona.
    Escrita em DB é assíncrona (asyncio.to_thread) — não bloqueia o loop.
    Tipos: status, carregado, dispensado, erro, limpeza_ok, telemetria

    Divisão de responsabilidade (ver ANALISE_ARQUITETURAL.md):
      - o simulador é fonte de verdade sobre HARDWARE/ESTOQUE
        (medicamento, sku, categoria, quantidade);
      - o orquestrador é fonte de verdade sobre FLUXO
        (status da etapa e os_id em execução).
    O evento "status" é telemetria periódica (a cada 15s, para todos os slots)
    e só pode tocar na primeira categoria. Os eventos de transição — carregado,
    dispensado, limpeza_ok, erro — é que movem o fluxo.
    """
    tipo    = payload.get("tipo", "status")
    disp_id = payload.get("dispenser_id")
    os_id   = payload.get("os_id")

    if disp_id is None:
        return

    disp_key = str(disp_id)
    db_tasks = []  # (fn, args) a executar fora do lock

    # ── Atualiza estado em memória (síncrono, sem await) ──────────────────
    with _lock:
        if disp_key not in _estado["dispensers"]:
            return
        d = _estado["dispensers"][disp_key]

        if tipo == "status":
            # Telemetria periódica: só estoque. Escrever "status"/"os_id" aqui
            # desfazia, a cada 15s, o reset de fim de OS do orquestrador — o
            # slot voltava a "concluido" para sempre e a limpeza dava 409.
            med = payload.get("medicamento")
            qty = payload.get("quantidade", 0)
            d.update({
                "medicamento": med,
                "sku":         payload.get("sku"),
                "categoria":   payload.get("categoria"),
                "quantidade":  qty,
            })
            med_db = med if (qty or 0) > 0 else None
            cat_db = payload.get("categoria") if (qty or 0) > 0 else None
            # ultima_os_id vem do fluxo que o central conhece, não do payload.
            db_tasks.append((salvar_dispenser_estado,
                             (int(disp_id), qty or 0, d["os_id"], med_db, cat_db)))

        elif tipo == "carregado":
            med = payload.get("medicamento")
            qty = payload.get("quantidade_total", payload.get("quantidade", 0))
            d.update({
                "status":      "pronto",
                "medicamento": med,
                "quantidade":  qty,
                "os_id":       os_id,
            })
            db_tasks.append((salvar_dispenser_estado,
                             (int(disp_id), qty or 0, os_id,
                              med if (qty or 0) > 0 else None,
                              payload.get("categoria"))))

        elif tipo == "dispensado":
            qtd_disp  = payload.get("quantidade_dispensada", 0)
            qtd_alvo  = payload.get("quantidade_alvo", 0)
            residual  = payload.get("quantidade_residual", 0)
            # Dispenser envia 'falha_mecanica' (bool) — 'validado' é derivado
            falha_mec = payload.get("falha_mecanica", False)
            validado  = not falha_mec and (qtd_disp >= qtd_alvo)
            falha     = payload.get("motivo_falha")
            med       = d.get("medicamento", payload.get("medicamento", ""))

            d.update({
                "status":                "concluido" if qtd_disp >= qtd_alvo else "dispensando",
                "quantidade_dispensada": qtd_disp,
                "quantidade_residual":   residual,
                "quantidade":            residual,
            })
            if residual == 0:
                d["medicamento"] = None
                d["sku"]         = None
                d["categoria"]   = None

            med_db = d["medicamento"] if residual > 0 else None
            cat_db = d.get("categoria") if residual > 0 else None
            db_tasks.append((salvar_dispensa,
                             (os_id, disp_id, med, qtd_disp, qtd_alvo, validado, falha)))
            db_tasks.append((salvar_dispenser_estado,
                             (int(disp_id), residual, os_id, med_db, cat_db)))
            if os_id:
                status_item = "concluido" if qtd_disp >= qtd_alvo else "em_andamento"
                db_tasks.append((atualizar_item_os, (os_id, disp_id, qtd_disp, status_item)))
            if not validado:
                db_tasks.append((salvar_alarme,
                                 (f"dispenser_{disp_id}", "falha_validacao",
                                  falha or f"Remédio rejeitado D{disp_id}")))

        elif tipo == "erro":
            d["status"] = "erro"
            descricao = payload.get("descricao", f"Erro no dispenser {disp_id}")
            db_tasks.append((salvar_alarme,
                             (f"dispenser_{disp_id}",
                              payload.get("codigo_erro", "erro"), descricao)))
            _log("alarme", f"ERRO D{disp_id}: {descricao}")

        elif tipo == "limpeza_ok":
            d.update({"status": "limpo", "medicamento": None, "sku": None,
                      "categoria": None, "quantidade": 0, "os_id": None})
            db_tasks.append((limpar_dispenser_estado, (int(disp_id),)))
            _log("limpeza_ok", f"D{disp_id} limpo.")

        elif tipo == "telemetria":
            componente = payload.get("componente", f"dispenser_{disp_id}")
            valor      = payload.get("valor_c", payload.get("valor", 0.0))
            unidade    = payload.get("unidade", "°C")
            tipo_leit  = payload.get("tipo_leitura", "temperatura")
            db_tasks.append((salvar_leitura_sensor,
                             (componente, tipo_leit, valor, unidade)))

    # ── Escreve no DB (async, fora do lock) ───────────────────────────────
    for fn, args in db_tasks:
        await _db(fn, *args)

    # Alarme recém-gravado entra na conta agora; sem alarme, a releitura só
    # acontece se o cache de _ALARMES_TTL_S tiver expirado. Vale para os quatro
    # handlers — o valor sai daqui direto para o `_broadcast_estado()` do fim.
    await _atualizar_alarmes_ativos(forcar=_abriu_alarme(db_tasks))

    # ── Notifica orquestrador (no event loop, seguro) ─────────────────────
    if os_id:
        if tipo == "carregado":
            orch.notificar_evento(f"{os_id}:carregado:{disp_id}", payload)
        elif tipo == "dispensado":
            orch.notificar_evento(f"{os_id}:dispensado:{disp_id}", payload)
        elif tipo == "erro":
            orch.notificar_evento(f"{os_id}:carregado:{disp_id}", {**payload, "tipo": "erro"})
            orch.notificar_evento(f"{os_id}:dispensado:{disp_id}", {**payload, "tipo": "erro"})

    # Limpeza é operação de slot, não de OS: o payload de "limpeza_ok" não traz
    # os_id, então a chave é só o dispenser. O "erro" também entra aqui — é
    # como o simulador recusa a limpeza (codigo_erro=limpeza_em_operacao) e o
    # orquestrador não pode ficar esperando uma confirmação que não vem.
    if tipo == "limpeza_ok":
        orch.notificar_evento(f"limpeza:{disp_id}", payload)
    elif tipo == "erro":
        orch.notificar_evento(f"limpeza:{disp_id}", {**payload, "tipo": "erro"})

    _log(f"disp_{tipo}", f"D{disp_id}: {tipo}", payload)
    # "status" e "telemetria" são os 12 eventos periódicos a cada 15s; as
    # transições (carregado, dispensado, erro, limpeza_ok) saem na hora.
    _broadcast_estado(prioritario=tipo not in _TIPOS_ALTA_FREQUENCIA)


async def _handle_evento_cnc(payload: dict):
    """
    Processa eventos do cnc-adapter.
    Estado atualizado síncronamente sob _lock; DB writes via asyncio.to_thread.
    Tipos: movendo, posicionado, concluido, erro, retornando, telemetria
    """
    tipo       = payload.get("tipo", "movendo")
    os_id      = payload.get("os_id")
    disp_alvo  = payload.get("dispenser_alvo")
    pos_x      = payload.get("posicao_x", 0.0)
    pos_y      = payload.get("posicao_y", 0.0)
    ciclo      = payload.get("ciclo_atual", 0)
    total      = payload.get("total_ciclos", 0)
    db_tasks   = []

    with _lock:
        _estado["cnc"].update({
            "status":         tipo,
            "os_id":          os_id,
            "dispenser_alvo": disp_alvo,
            "posicao_x":      pos_x,
            "posicao_y":      pos_y,
            "ciclo_atual":    ciclo,
            "total_ciclos":   total,
        })
        if _estado["os_ativa"] and os_id:
            if _estado["os_ativa"].get("os_id") == os_id and tipo == "concluido":
                _estado["os_ativa"]["status"] = "concluida"

        # Só TRANSIÇÃO vira linha. "movendo" chega a cada 0.5s durante todo o
        # movimento — eram centenas de linhas por OS para descrever uma
        # trajetória que o dashboard já mostra ao vivo e que ninguém consulta
        # depois. Quem quiser rastro grava amostrado (CNC_AMOSTRAGEM_MOVENDO).
        if tipo in ("posicionado", "concluido", "erro"):
            db_tasks.append((salvar_cnc_evento,
                             (os_id, tipo, disp_alvo, pos_x, pos_y, ciclo, total)))
        elif tipo == "movendo" and _amostrar_movendo():
            db_tasks.append((salvar_cnc_evento,
                             (os_id, tipo, disp_alvo, pos_x, pos_y, ciclo, total)))

        if tipo == "erro":
            descricao = payload.get("descricao", "Erro desconhecido na CNC")
            db_tasks.append((salvar_alarme,
                             ("cnc", payload.get("codigo_erro", "erro_cnc"), descricao)))
            _log("alarme", f"ERRO CNC: {descricao}")

        elif tipo == "telemetria":
            componente = payload.get("componente", "cnc")
            valor      = payload.get("valor", 0.0)
            unidade    = payload.get("unidade", "°C")
            tipo_leit  = payload.get("tipo_leitura", "temperatura")
            db_tasks.append((salvar_leitura_sensor,
                             (componente, tipo_leit, valor, unidade)))

    # DB fora do lock
    for fn, args in db_tasks:
        await _db(fn, *args)

    await _atualizar_alarmes_ativos(forcar=_abriu_alarme(db_tasks))

    # Notifica orquestrador
    if os_id and disp_alvo is not None:
        if tipo == "posicionado":
            orch.notificar_evento(f"{os_id}:posicionado:{disp_alvo}", payload)
        elif tipo == "erro":
            orch.notificar_evento(f"{os_id}:posicionado:{disp_alvo}",
                                  {**payload, "tipo": "erro"})

    _log(f"cnc_{tipo}", f"CNC {tipo} | D{disp_alvo} | ({pos_x:.1f},{pos_y:.1f})")
    _broadcast_estado(prioritario=tipo not in _TIPOS_ALTA_FREQUENCIA)


async def _handle_evento_visao(payload: dict):
    """
    Processa eventos do vision-adapter.
    Tipos (câmera dispenser): leitura_dispenser_ok, leitura_dispenser_falha, leitura_dispenser_divergencia
    Tipos (câmera mesa):      leitura_mesa_ok, leitura_mesa_falha, leitura_mesa_divergencia
    Tipo  (telemetria):       telemetria (temperatura dos componentes de câmera)

    Estado atualizado síncronamente sob _lock.
    Alarmes gerados assincronamente via asyncio.to_thread.
    Orquestrador notificado via notificar_evento (call_soon_threadsafe).
    """
    tipo    = payload.get("tipo", "")
    camera  = payload.get("camera", "")
    slot_id = payload.get("slot_id")
    os_id   = payload.get("os_id")
    db_tasks = []

    with _lock:
        # ── Câmera dos Dispensers ──────────────────────────────────────────
        if camera == "dispenser" and tipo in (
            "leitura_dispenser_ok",
            "leitura_dispenser_falha",
            "leitura_dispenser_divergencia",
        ):
            _estado["visao"]["camera_dispenser"].update({
                "ultima_leitura": tipo,
                "slot_id":        slot_id,
                "match_sku":      payload.get("match_sku"),
                "confianca":      payload.get("confianca"),
                "ts":             payload.get("ts"),
            })

            # Persiste leitura no histórico (todos os tipos de câmera dispenser)
            db_tasks.append((salvar_leitura_visao, (
                os_id, "dispenser", slot_id, tipo,
                payload.get("sku_esperado"), payload.get("sku_lido"),
                payload.get("match_sku"), payload.get("confianca"),
                None, None, payload.get("motivo"),
            )))

            if tipo == "leitura_dispenser_falha":
                descricao = (f"Falha câmera dispenser slot {slot_id}: "
                             f"{payload.get('motivo', 'desconhecido')}")
                db_tasks.append((salvar_alarme,
                                 (f"camera_dispenser_{slot_id}",
                                  "falha_leitura_dispenser", descricao)))
                _log("alarme", descricao)

            elif tipo == "leitura_dispenser_divergencia":
                descricao = (f"SKU incorreto slot {slot_id}: "
                             f"esperado={payload.get('sku_esperado','?')} "
                             f"lido={payload.get('sku_lido','?')}")
                db_tasks.append((salvar_alarme,
                                 (f"camera_dispenser_{slot_id}",
                                  "divergencia_sku", descricao)))
                _log("alarme", descricao)

        # ── Câmera da Mesa ─────────────────────────────────────────────────
        elif camera == "mesa" and tipo in (
            "leitura_mesa_ok",
            "leitura_mesa_falha",
            "leitura_mesa_divergencia",
        ):
            _estado["visao"]["camera_mesa"].update({
                "ultima_leitura":       tipo,
                "slot_id":              slot_id,
                "quantidade_detectada": payload.get("quantidade_detectada"),
                "quantidade_esperada":  payload.get("quantidade_esperada"),
                "confianca":            payload.get("confianca"),
                "ts":                   payload.get("ts"),
            })

            # Persiste leitura no histórico (todos os tipos de câmera mesa)
            db_tasks.append((salvar_leitura_visao, (
                os_id, "mesa", slot_id, tipo,
                None, None, None, payload.get("confianca"),
                payload.get("quantidade_esperada"), payload.get("quantidade_detectada"),
                payload.get("motivo"),
            )))

            if tipo == "leitura_mesa_falha":
                descricao = (f"Câmera mesa não detectou produto slot {slot_id}: "
                             f"{payload.get('motivo', 'desconhecido')}")
                db_tasks.append((salvar_alarme,
                                 (f"camera_mesa_{slot_id}",
                                  "falha_deteccao_mesa", descricao)))
                _log("alarme", descricao)

            elif tipo == "leitura_mesa_divergencia":
                det = payload.get("quantidade_detectada", 0)
                esp = payload.get("quantidade_esperada", 0)
                descricao = (f"Contagem incorreta slot {slot_id}: "
                             f"esperado={esp} detectado={det} "
                             f"(Δ={det - esp:+d})")
                db_tasks.append((salvar_alarme,
                                 (f"camera_mesa_{slot_id}",
                                  "divergencia_contagem", descricao)))
                _log("alarme", descricao)

        # ── Telemetria das câmeras ──────────────────────────────────────────
        elif tipo == "telemetria":
            componente = payload.get("componente", "camera_sistema")
            valor      = payload.get("valor", 0.0)
            unidade    = payload.get("unidade", "°C")
            tipo_leit  = payload.get("tipo_leitura", "temperatura")
            db_tasks.append((salvar_leitura_sensor,
                             (componente, tipo_leit, valor, unidade)))

    # ── DB fora do lock ────────────────────────────────────────────────────
    for fn, args in db_tasks:
        await _db(fn, *args)

    await _atualizar_alarmes_ativos(forcar=_abriu_alarme(db_tasks))

    # ── Notifica orquestrador ──────────────────────────────────────────────
    if os_id and slot_id is not None:
        # Dispenser: qualquer resultado (ok, falha, divergência) desbloqueia o await
        if tipo.startswith("leitura_dispenser_"):
            orch.notificar_evento(f"{os_id}:visao_dispenser:{slot_id}", payload)
        # Mesa: idem
        elif tipo.startswith("leitura_mesa_"):
            orch.notificar_evento(f"{os_id}:visao_mesa:{slot_id}", payload)

    _log(f"visao_{tipo}", f"cam={camera} slot={slot_id}", payload)
    _broadcast_estado(prioritario=tipo not in _TIPOS_ALTA_FREQUENCIA)


async def _handle_evento_peso(payload: dict):
    """
    Processa eventos do weight-adapter (balança HX711).
    Tipos: tara_ok, peso_ok, peso_divergencia, erro_sensor, telemetria
    """
    tipo    = payload.get("tipo", "")
    slot_id = payload.get("slot_id")
    os_id   = payload.get("os_id")
    db_tasks = []

    with _lock:
        if tipo in ("peso_ok", "peso_divergencia", "tara_ok"):
            _estado["peso"].update({
                "ultima_leitura":   tipo,
                "slot_id":          slot_id,
                "peso_medido_g":    payload.get("peso_medido_g"),
                "peso_esperado_g":  payload.get("peso_esperado_g"),
                "desvio_pct":       payload.get("desvio_pct"),
                "ts":               payload.get("ts"),
            })

        if tipo == "peso_divergencia":
            descricao = (
                f"Divergência de peso slot {slot_id}: "
                f"esperado={(payload.get('peso_esperado_g') or 0):.1f}g "
                f"medido={(payload.get('peso_medido_g') or 0):.1f}g "
                f"(desvio={(payload.get('desvio_pct') or 0):.1f}%)"
            )
            db_tasks.append((salvar_alarme,
                             (f"balanca_{slot_id}", "divergencia_peso", descricao)))
            _log("alarme", descricao)

        elif tipo == "erro_sensor":
            descricao = f"Sensor HX711 falhou: {payload.get('descricao', '')}"
            db_tasks.append((salvar_alarme,
                             ("balanca", "erro_sensor_peso", descricao)))

        elif tipo == "telemetria":
            componente = payload.get("componente", "hx711_balanca_mesa")
            db_tasks.append((salvar_leitura_sensor,
                             (componente, "temperatura", payload.get("temperatura_c", 0), "°C")))

    for fn, args in db_tasks:
        await _db(fn, *args)

    await _atualizar_alarmes_ativos(forcar=_abriu_alarme(db_tasks))

    # Notifica orquestrador (para tara e pesagem de slots)
    if os_id:
        if tipo == "tara_ok":
            orch.notificar_evento(f"{os_id}:tara", payload)
        elif tipo in ("peso_ok", "peso_divergencia", "erro_sensor") and slot_id is not None:
            orch.notificar_evento(f"{os_id}:peso:{slot_id}", payload)

    _log(f"peso_{tipo}", f"slot={slot_id}", payload)
    _broadcast_estado(prioritario=tipo not in _TIPOS_ALTA_FREQUENCIA)


# ── Expurgo periódico do histórico ────────────────────────────────────────────

async def _loop_expurgo():
    """Apaga `cnc_eventos` e `leituras_sensores` além da retenção.

    Roda no startup (banco herdado de versão sem expurgo pode chegar grande) e
    depois a cada `EXPURGO_INTERVALO_HORAS`. Falha de expurgo é registrada e
    ignorada: é manutenção, não pode derrubar a operação.
    """
    intervalo = max(settings.EXPURGO_INTERVALO_HORAS, 0.01) * 3600
    while True:
        try:
            removidos = await asyncio.to_thread(
                expurgar_dados_antigos, settings.RETENCAO_DIAS
            )
            if any(removidos.values()):
                _log("expurgo", f"Histórico expurgado (> {settings.RETENCAO_DIAS}d): "
                                + ", ".join(f"{t}={n}" for t, n in removidos.items()))
        except Exception as exc:
            logger.warning("[DB] Expurgo falhou: %s", exc)
        await asyncio.sleep(intervalo)


# ── FastAPI Lifespan ───────────────────────────────────────────────────────────
@asynccontextmanager
async def lifespan(app: FastAPI):
    global _loop
    # FIX: get_running_loop() em vez de get_event_loop() (deprecado em Python ≥3.10)
    _loop = asyncio.get_running_loop()

    await asyncio.to_thread(init_db)

    # Alarmes abertos sobrevivem ao processo: sem esta leitura o badge nasce em
    # 0 depois de todo restart, com o banco cheio de alarmes por resolver.
    await _atualizar_alarmes_ativos(forcar=True)

    # Injeta loop no orquestrador para notificar_evento thread-safe
    orch.inicializar(_estado, _lock, _broadcast_estado, _loop)

    # Segredo fraco derruba o boot (fora de APSEN_ENV=dev). O valor default é
    # público neste repositório: com ele qualquer um forja um JWT role=admin.
    validar_secret_key(settings.SECRET_KEY, settings.APSEN_ENV)

    task_broadcast = asyncio.create_task(_broadcast_worker())
    task_flusher   = asyncio.create_task(_broadcast_flusher())
    task_orch      = asyncio.create_task(orch.loop_orquestrador())
    task_expurgo   = asyncio.create_task(_loop_expurgo())

    logger.info("Computador Central APSEN v3.1 iniciado.")
    yield

    for tarefa in (task_broadcast, task_flusher, task_orch, task_expurgo):
        tarefa.cancel()
    await orch.encerrar()


app = FastAPI(title="APSEN Computador Central v3.1", lifespan=lifespan)
# Só dashboard e IHM falam com o central pelo navegador. `*` num serviço
# autenticado deixa qualquer página aberta no browser do técnico disparar
# requisição em nome dele.
app.add_middleware(CORSMiddleware, allow_origins=settings.CORS_ORIGINS,
                   allow_credentials=True,
                   allow_methods=["*"], allow_headers=["*"])

_bearer = HTTPBearer(auto_error=False)


# ── Revalidação do usuário a cada requisição ──────────────────────────────────
# O JWT era a palavra final: `_get_tecnico` decodificava e pronto. Um técnico
# desativado seguia com acesso total até o token expirar — 8 horas —, e uma
# role rebaixada continuava valendo pelo mesmo tempo. Como o token é assinado,
# nem dá para "editar" o que já foi emitido: a fonte de verdade tem que ser o
# banco, consultado a cada requisição.
#
# Cache de AUTH_CACHE_TTL_S (30s) para não virar uma query por request: é o
# atraso máximo entre desativar alguém e o acesso cair. As mutações que o
# próprio central conhece (desativar, ativar, trocar role) invalidam a entrada
# na hora, então na prática o atraso só existe para mudança feita fora da API.
_cache_usuarios: dict[str, tuple[float, dict]] = {}
_cache_usuarios_lock = threading.Lock()


def _invalidar_cache_usuario(username: str) -> None:
    with _cache_usuarios_lock:
        _cache_usuarios.pop(username, None)


def _usuario_valido(username: str) -> Optional[dict]:
    """Usuário ATIVO do banco, ou None. `get_usuario` já filtra `ativo=1`."""
    agora = time.monotonic()
    with _cache_usuarios_lock:
        entrada = _cache_usuarios.get(username)
        if entrada and (agora - entrada[0]) < settings.AUTH_CACHE_TTL_S:
            return entrada[1]

    try:
        usuario = get_usuario(username)
    except Exception as exc:
        # Banco fora do ar não pode virar acesso liberado.
        logger.error("[AUTH] Falha ao revalidar '%s': %s", username, exc)
        return None

    with _cache_usuarios_lock:
        _cache_usuarios[username] = (time.monotonic(), usuario)
    return usuario


def _get_tecnico(creds: Optional[HTTPAuthorizationCredentials] = Depends(_bearer)):
    if not creds:
        raise HTTPException(status.HTTP_401_UNAUTHORIZED, "Token ausente")
    payload = decodificar_token(creds.credentials)
    if not payload:
        raise HTTPException(status.HTTP_401_UNAUTHORIZED, "Token inválido")

    usuario = _usuario_valido(payload.get("sub", ""))
    if not usuario:
        raise HTTPException(status.HTTP_401_UNAUTHORIZED,
                            "Usuário inativo ou inexistente")
    # A role vem do BANCO, não do token: ela pode ter mudado desde a emissão,
    # e um token forjado traria a role que o portador quisesse.
    return {**payload, "role": usuario.get("role", "manutencao")}


def _get_admin(user=Depends(_get_tecnico)):
    if user.get("role") != "admin":
        raise HTTPException(status.HTTP_403_FORBIDDEN, "Requer perfil admin")
    return user


# ── Pydantic Models ────────────────────────────────────────────────────────────
class LoginReq(BaseModel):
    username: str
    senha: str


class ManutencaoReq(BaseModel):
    tipo: str
    componente: str
    descricao: str


class StatusOSReq(BaseModel):
    status: str


class UsuarioReq(BaseModel):
    username: str
    senha: str
    nome_completo: str
    role: str = "manutencao"


class UsuarioUpdateReq(BaseModel):
    nome_completo: Optional[str] = None
    role: Optional[str] = None
    nova_senha: Optional[str] = None


class NovaOSReq(BaseModel):
    os_id: str
    descricao: str = ""
    categoria: str = ""
    medicamentos: list


class EventoDispenserReq(BaseModel):
    tipo: str
    dispenser_id: Optional[int] = None
    os_id: Optional[str] = None

    model_config = {"extra": "allow"}


class EventoCNCReq(BaseModel):
    tipo: str
    os_id: Optional[str] = None
    dispenser_alvo: Optional[int] = None

    model_config = {"extra": "allow"}


class EventoVisionReq(BaseModel):
    tipo: str
    camera: str                     # "dispenser" | "mesa" | "sistema"
    slot_id: Optional[int] = None
    os_id: Optional[str] = None

    model_config = {"extra": "allow"}


class EventoPesoReq(BaseModel):
    tipo: str                       # tara_ok | peso_ok | peso_divergencia | erro_sensor | telemetria
    os_id: Optional[str] = None
    slot_id: Optional[int] = None

    model_config = {"extra": "allow"}


# ══════════════════════════════════════════════════════════════════════════════
# ENDPOINTS — RECEBIMENTO (adapters e order-generator)
# ══════════════════════════════════════════════════════════════════════════════

@app.post("/api/v1/ordens")
async def receber_ordem(req: NovaOSReq):
    """Recebe nova OS do order-generator.

    **Nenhuma OS é enfileirada sem linha no banco.** A fila é o que dispensa
    medicamento; o banco é o que registra o que foi dispensado. Aceitar uma
    sem a outra produz os dois piores resultados do sistema:

      - `salvar_ordem` devolvendo False (INSERT IGNORE não inseriu) significa
        OS já registrada. Enfileirar de novo processa a MESMA OS duas vezes —
        dose dobrada no leito. Vira 409 `os_duplicada`, o contrato da
        ANALISE_ARQUITETURAL §4.1.
      - `salvar_ordem` levantando significa banco fora do ar. Antes isso só era
        logado e a OS seguia para a fila: os medicamentos sairiam do dispenser
        sem linha em `ordens`/`os_itens`, então `atualizar_item_os` e
        `atualizar_status_ordem` não achariam o que atualizar e o relatório
        (`GET /os/{os_id}`, CSV/XLSX) sairia vazio — dispensa sem rastro, que é
        justamente o que um sistema de medicação não pode fazer. Vira 503, e a
        OS não entra na fila.

    Os corpos seguem a forma `{"erro": ..., "os_id": ...}` do contrato, e por
    isso são `JSONResponse` — `HTTPException` embrulharia tudo em `detail`.
    Em todos os casos o order-generator apenas loga e segue para a próxima OS
    no ciclo seguinte (`_enviar_os` devolve False e ninguém retenta), então não
    há laço de reenvio nem processo derrubado.

    A terceira recusa é de fila cheia (429): o gerador posta mais rápido do que
    a planta processa, e sob trava do Triple Check o orquestrador para até um
    humano liberar. Ela vem ANTES de `salvar_ordem` — OS recusada não gera
    linha no banco, senão o "aguardando" órfão viraria a OS ativa aos olhos de
    `get_ordem_ativa`.
    """
    os_id        = req.os_id
    medicamentos = req.medicamentos

    if not os_id or not medicamentos:
        raise HTTPException(400, "os_id e medicamentos são obrigatórios")

    if not orch.ha_vaga_na_fila():
        fila = orch.fila_status()
        logger.warning("[API] OS %s recusada — fila cheia (%d/%d).",
                       os_id, fila["tamanho"], fila["capacidade"])
        _log("os_recusada", f"OS {os_id} recusada: fila cheia "
                            f"({fila['tamanho']}/{fila['capacidade']})")
        return JSONResponse(
            status_code=429,
            content={
                "erro":        "fila_cheia",
                "os_id":       os_id,
                "fila":        fila,
                "mensagem":    (f"Fila de OS no limite ({fila['tamanho']}/"
                                f"{fila['capacidade']}). Tente novamente mais tarde."),
            },
            headers={"Retry-After": str(int(settings.TIMEOUT_DISPENSA))},
        )

    try:
        inserida = await asyncio.to_thread(
            salvar_ordem, os_id, req.descricao, medicamentos, req.model_dump()
        )
    except Exception as exc:
        logger.error("[DB] Erro ao salvar OS %s: %s — OS recusada.", os_id, exc)
        _log("os_recusada", f"OS {os_id} recusada: persistência indisponível")
        return JSONResponse(
            status_code=503,
            content={"erro": "persistencia_indisponivel", "os_id": os_id},
        )

    if not inserida:
        logger.warning("[API] OS %s já registrada — recusada (duplicada).", os_id)
        _log("os_duplicada", f"OS {os_id} recusada: já registrada")
        return JSONResponse(
            status_code=409,
            content={"erro": "os_duplicada", "os_id": os_id},
        )

    with _lock:
        posicao_fila = len(_estado["fila_os"]) + (1 if _estado["os_ativa"] else 0)

    if not await orch.enfileirar_os(req.model_dump()):
        # Corrida perdida: a vaga conferida acima sumiu entre a checagem e o
        # enfileiramento. A OS já foi persistida, e uma linha "aguardando" que
        # ninguém vai processar seria devolvida por `get_ordem_ativa` como a OS
        # ativa — daí o fechamento em "cancelada".
        try:
            await asyncio.to_thread(atualizar_status_ordem, os_id, "cancelada")
        except Exception as exc:
            logger.error("[DB] cancelar OS %s não enfileirada: %s", os_id, exc)
        fila = orch.fila_status()
        return JSONResponse(
            status_code=429,
            content={"erro": "fila_cheia", "os_id": os_id, "fila": fila,
                     "mensagem": "Fila de OS ficou cheia durante o registro."},
            headers={"Retry-After": str(int(settings.TIMEOUT_DISPENSA))},
        )

    _log("os_nova", f"OS {os_id} recebida — {len(medicamentos)} medicamento(s)")
    logger.info("[API] OS %s recebida.", os_id)
    _broadcast_estado()

    return {"aceita": True, "os_id": os_id, "posicao_fila": posicao_fila + 1}


@app.post("/api/v1/eventos/dispenser")
async def evento_dispenser(req: EventoDispenserReq):
    """Recebe eventos normalizados do dispenser-adapter."""
    await _handle_evento_dispenser(req.model_dump())
    return {"ok": True}


@app.post("/api/v1/eventos/cnc")
async def evento_cnc(req: EventoCNCReq):
    """Recebe eventos normalizados do cnc-adapter."""
    await _handle_evento_cnc(req.model_dump())
    return {"ok": True}


@app.post("/api/v1/eventos/visao")
async def evento_visao(req: EventoVisionReq):
    """Recebe resultados de captura do vision-adapter (câmera dispenser e câmera mesa)."""
    await _handle_evento_visao(req.model_dump())
    return {"ok": True}


@app.post("/api/v1/eventos/peso")
async def evento_peso(req: EventoPesoReq):
    """Recebe leituras de pesagem do weight-adapter (balança HX711)."""
    await _handle_evento_peso(req.model_dump())
    return {"ok": True}


# ══════════════════════════════════════════════════════════════════════════════
# ENDPOINTS — DASHBOARD (sem autenticação, read-only)
# ══════════════════════════════════════════════════════════════════════════════

@app.get("/ping")
def ping():
    return {"status": "ok", "service": "apsen-central-computer"}


@app.get("/estado")
def get_estado():
    # Endpoint síncrono (roda em threadpool), então a releitura pode bloquear.
    # Sujeita ao mesmo cache dos handlers: o polling do dashboard não vira uma
    # query por request.
    _contar_alarmes_ativos()
    with _lock:
        # deepcopy, não dict(): a cópia rasa devolve os MESMOS dicionários
        # aninhados (dispensers, cnc, visao, peso, trava). A serialização
        # acontece depois do `with`, já sem o lock, e um evento chegando nesse
        # intervalo muda o dicionário durante a iteração do serializador
        # ("dictionary changed size during iteration"). O mesmo padrão de
        # `_broadcast_estado`.
        return copy.deepcopy(_estado)


@app.get("/api/v1/fila")
def get_fila():
    """Ocupação da fila de OS — endpoint de backpressure do order-generator.

    O `/estado` já traz `fila_tamanho`, mas serve o snapshot inteiro (6 slots,
    CNC, visão, peso, trava) e ainda passa pelo contador de alarmes. Quem só
    precisa saber se cabe mais uma OS não deveria pagar por isso a cada ciclo —
    daí este endpoint de dois números. `fila_capacidade` continua no `/estado`
    para o dashboard mostrar a escala junto do resto.
    """
    return orch.fila_status()


@app.get("/os/ativa")
def os_ativa():
    ordem = get_ordem_ativa()
    if not ordem:
        return {"os_ativa": None}
    return {"os_ativa": ordem}


@app.get("/os/historico")
def os_historico(limite: int = 50):
    return get_historico_ordens(limite)


@app.get("/os/{os_id}")
def os_detalhe(os_id: str):
    ordem = get_ordem_por_id(os_id)
    if not ordem:
        raise HTTPException(404, "OS não encontrada")
    return ordem


@app.get("/api/v1/relatorio/os/{os_id}")
async def relatorio_os(
    os_id: str,
    formato: str = "csv",
    creds: Optional[HTTPAuthorizationCredentials] = Depends(_bearer),
):
    """
    Exporta relatório completo de uma OS em CSV ou XLSX.

    Autenticação **só** por `Authorization: Bearer <jwt>`. O `?token=` que
    existia aqui servia para o navegador baixar direto de uma âncora da IHM, o
    que punha o JWT no histórico do navegador, no `Referer` e no log de acesso
    deste serviço. Hoje quem baixa é o processo da IHM (server-side, dentro da
    rede Docker) e devolve os bytes ao operador pelo `dcc.Download` — nenhum
    cliente precisa mais de token na URL.
    """
    raw_token = creds.credentials if creds else None
    if not raw_token or not decodificar_token(raw_token):
        raise HTTPException(status.HTTP_401_UNAUTHORIZED, "Token ausente ou inválido")
    # Buscar dados em paralelo
    os_data, dispensas_data, visao_data, alarmes_data = await asyncio.gather(
        asyncio.to_thread(get_ordem_por_id, os_id),
        asyncio.to_thread(get_dispensas, os_id, 200),
        asyncio.to_thread(get_historico_visao, os_id, 200),
        asyncio.to_thread(get_alarmes_por_os, os_id),
    )
    if not os_data:
        raise HTTPException(status_code=404, detail=f"OS '{os_id}' não encontrada.")

    if formato.lower() == "xlsx":
        return await _gerar_xlsx(os_id, os_data, dispensas_data, visao_data, alarmes_data)
    return _gerar_csv(os_id, os_data, dispensas_data, visao_data, alarmes_data)


def _gerar_csv(os_id, os_data, dispensas_data, visao_data, alarmes_data) -> StreamingResponse:
    """Gera CSV com múltiplas seções separadas por linha em branco."""
    buf = io.StringIO()
    writer = csv.writer(buf)

    # Cabeçalho OS
    writer.writerow(["=== ORDEM DE SAÍDA ==="])
    writer.writerow(["OS ID", "Status", "Categoria", "Descrição", "Criado em", "Atualizado em"])
    writer.writerow([
        os_data.get("os_id", ""), os_data.get("status", ""),
        os_data.get("categoria", ""), os_data.get("descricao", ""),
        str(os_data.get("criado_em", ""))[:19],
        str(os_data.get("atualizado_em", ""))[:19],
    ])
    writer.writerow([])

    # Dispensas
    writer.writerow(["=== DISPENSAS ==="])
    writer.writerow(["Dispenser", "Medicamento", "Qtd Dispensada", "Qtd Alvo", "Validado", "Horário"])
    for d in (dispensas_data or []):
        writer.writerow([
            f"D{d.get('dispenser_id','')}",
            d.get("medicamento", ""),
            d.get("quantidade_dispensada", ""),
            d.get("quantidade_alvo", ""),
            "SIM" if d.get("validado") else "NÃO",
            str(d.get("ts", ""))[:19],
        ])
    writer.writerow([])

    # Visão computacional
    writer.writerow(["=== LEITURAS DE VISÃO COMPUTACIONAL ==="])
    writer.writerow(["Câmera", "Slot", "Tipo", "SKU Esperado", "SKU Lido", "Match", "Confiança", "Qtd Det.", "Qtd Esp.", "Horário"])
    for v in (visao_data or []):
        writer.writerow([
            v.get("camera", ""), f"D{v.get('slot_id','')}",
            v.get("tipo", ""),
            v.get("sku_esperado", ""), v.get("sku_lido", ""),
            "SIM" if v.get("match_sku") == 1 else ("NÃO" if v.get("match_sku") == 0 else "—"),
            f"{(v.get('confianca') or 0)*100:.1f}%",
            v.get("qtd_detectada", ""), v.get("qtd_esperada", ""),
            str(v.get("criado_em", ""))[:19],
        ])
    writer.writerow([])

    # Alarmes
    writer.writerow(["=== ALARMES ==="])
    writer.writerow(["Fonte", "Tipo", "Descrição", "Horário"])
    for a in (alarmes_data or []):
        writer.writerow([
            a.get("fonte", ""), a.get("tipo", ""),
            a.get("descricao", ""),
            str(a.get("ts", ""))[:19],
        ])

    buf.seek(0)
    filename = f"relatorio_{os_id}.csv"
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


async def _gerar_xlsx(os_id, os_data, dispensas_data, visao_data, alarmes_data) -> StreamingResponse:
    """Gera XLSX com múltiplas abas — requer openpyxl."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(
            status_code=501,
            detail="openpyxl não instalado no servidor. Use formato=csv.",
        )

    wb = openpyxl.Workbook()

    def _header_style(ws, row, cols):
        fill = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
        font = Font(bold=True, color="FFFFFF")
        for col, val in enumerate(cols, start=1):
            c = ws.cell(row=row, column=col, value=val)
            c.fill = fill
            c.font = font
            c.alignment = Alignment(horizontal="center")

    # Aba: OS
    ws_os = wb.active
    ws_os.title = "OS"
    _header_style(ws_os, 1, ["OS ID", "Status", "Categoria", "Descrição", "Criado em", "Atualizado em"])
    ws_os.append([
        os_data.get("os_id", ""), os_data.get("status", ""),
        os_data.get("categoria", ""), os_data.get("descricao", ""),
        str(os_data.get("criado_em", ""))[:19], str(os_data.get("atualizado_em", ""))[:19],
    ])

    # Aba: Dispensas
    ws_d = wb.create_sheet("Dispensas")
    _header_style(ws_d, 1, ["Dispenser", "Medicamento", "Qtd Dispensada", "Qtd Alvo", "Validado", "Horário"])
    for d in (dispensas_data or []):
        ws_d.append([
            f"D{d.get('dispenser_id','')}", d.get("medicamento", ""),
            d.get("quantidade_dispensada", ""), d.get("quantidade_alvo", ""),
            "SIM" if d.get("validado") else "NÃO",
            str(d.get("ts", ""))[:19],
        ])

    # Aba: Visão
    ws_v = wb.create_sheet("Visão CV")
    _header_style(ws_v, 1, ["Câmera", "Slot", "Tipo", "SKU Esp.", "SKU Lido", "Match", "Conf.", "Qtd Det.", "Qtd Esp.", "Horário"])
    for v in (visao_data or []):
        ws_v.append([
            v.get("camera", ""), f"D{v.get('slot_id','')}",
            v.get("tipo", ""), v.get("sku_esperado", ""), v.get("sku_lido", ""),
            "SIM" if v.get("match_sku") == 1 else ("NÃO" if v.get("match_sku") == 0 else "—"),
            f"{(v.get('confianca') or 0)*100:.1f}%",
            v.get("qtd_detectada", ""), v.get("qtd_esperada", ""),
            str(v.get("criado_em", ""))[:19],
        ])

    # Aba: Alarmes
    ws_a = wb.create_sheet("Alarmes")
    _header_style(ws_a, 1, ["Fonte", "Tipo", "Descrição", "Horário"])
    for a in (alarmes_data or []):
        ws_a.append([
            a.get("fonte", ""), a.get("tipo", ""),
            a.get("descricao", ""), str(a.get("ts", ""))[:19],
        ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"relatorio_{os_id}.xlsx"
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@app.get("/api/v1/visao/historico")
async def visao_historico(os_id: str = None, limite: int = 100):
    """Retorna histórico de leituras de visão computacional. Filtra por os_id se fornecido."""
    if limite > 500:
        limite = 500
    rows = await asyncio.to_thread(get_historico_visao, os_id, limite)
    return {"leituras": rows, "total": len(rows)}


@app.get("/medicamentos")
def get_medicamentos(categoria: str = None):
    return listar_medicamentos(categoria)


@app.get("/medicamentos/categorias")
def get_categorias():
    return listar_categorias()


@app.get("/dispensas")
def dispensas(os_id: str = None, limite: int = 100):
    if os_id:
        return get_dispensas(os_id, limite)
    return get_dispensas_recentes(limite)


@app.get("/dispensers/estado")
def dispensers_estado():
    return get_dispensers_estado()


@app.get("/cnc/historico")
def cnc_historico(limite: int = 50):
    return get_cnc_recentes(limite)


@app.get("/alarmes")
def alarmes(resolvido: bool = False, limite: int = 50):
    return get_alarmes(resolvido=resolvido, limite=limite)


@app.get("/log/eventos")
def log_eventos(limite: int = 50):
    return list(_log_eventos)[:limite]


# ══════════════════════════════════════════════════════════════════════════════
# ENDPOINTS — IHM MANUTENÇÃO (requer JWT)
# ══════════════════════════════════════════════════════════════════════════════

@app.post("/auth/login")
def login(req: LoginReq):
    user = get_usuario(req.username)
    if not user or not verificar_senha(req.senha, user["senha_hash"]):
        raise HTTPException(status.HTTP_401_UNAUTHORIZED, "Credenciais inválidas")
    role  = user.get("role", "manutencao")
    token = criar_token(user["username"], user["nome_completo"], role)
    return {
        "token":    token,
        "username": user["username"],
        "nome":     user["nome_completo"],
        "role":     role,
    }


@app.get("/auth/me")
def me(user=Depends(_get_tecnico)):
    return {"username": user["sub"], "nome": user["nome"], "role": user.get("role")}


@app.put("/ordens/{os_id}/status")
def alterar_status_os(os_id: str, req: StatusOSReq, user=Depends(_get_tecnico)):
    validos = {"aguardando", "em_andamento", "concluida", "erro", "cancelada"}
    if req.status not in validos:
        raise HTTPException(400, f"Status inválido. Use: {validos}")
    atualizar_status_ordem(os_id, req.status)
    _log("os_status", f"OS {os_id} → {req.status} por {user['sub']}")
    return {"ok": True, "os_id": os_id, "status": req.status}


@app.get("/manutencao/sensores")
def manut_sensores(user=Depends(_get_tecnico)):
    return get_ultimas_leituras()


@app.get("/manutencao/sensores/{componente}")
def manut_sensor_hist(componente: str, tipo: str = "temperatura", limite: int = 60,
                      user=Depends(_get_tecnico)):
    return get_historico_sensor(componente, tipo, limite)


@app.get("/manutencao/log")
def manut_log(limite: int = 100, user=Depends(_get_tecnico)):
    return get_log_manutencao(limite)


@app.post("/manutencao/log")
def manut_registrar(req: ManutencaoReq, user=Depends(_get_tecnico)):
    return salvar_manutencao(req.tipo, req.componente, req.descricao, user["sub"])


@app.get("/manutencao/alarmes")
def manut_alarmes(resolvido: bool = False, limite: int = 100, user=Depends(_get_tecnico)):
    return get_alarmes(resolvido=resolvido, limite=limite)


@app.put("/manutencao/alarmes/{alarme_id}/resolver")
def manut_resolver_alarme(alarme_id: int, user=Depends(_get_tecnico)):
    resolver_alarme(alarme_id)
    # Único caminho que ABAIXA o total — o que o contador manual nunca fez.
    # `forcar` porque o técnico está olhando o badge: esperar o TTL aqui é
    # exatamente o que faz o número parecer travado.
    total = _contar_alarmes_ativos(forcar=True)
    _log("alarme_resolvido", f"Alarme {alarme_id} resolvido por {user['sub']}")
    _broadcast_estado()
    return {"ok": True, "alarmes_ativos": total}


# ── Triple Check — trava de emergência ────────────────────────────────────────

@app.get("/api/v1/trava")
def get_trava():
    """Retorna estado atual da trava de Triple Check."""
    return orch.get_trava_estado()


@app.post("/api/v1/admin/liberar-trava")
async def liberar_trava(user=Depends(_get_admin)):
    """
    Libera a trava de Triple Check. Exige role admin ou supervisor.
    A OS retoma de onde parou após a liberação.
    """
    liberado = await asyncio.to_thread(orch.liberar_trava, user["sub"])
    if not liberado:
        raise HTTPException(status_code=409, detail="Nenhuma trava ativa no momento.")
    _log("trava", f"Trava liberada por {user['sub']}")
    with _lock:
        _estado["trava"] = {"ativa": False, "os_id": None, "slot_id": None, "motivo": ""}
    _broadcast_estado()
    return {"ok": True, "liberado_por": user["sub"]}


@app.post("/manutencao/dispensers/{dispenser_id}/limpar")
async def manut_limpar_dispenser(dispenser_id: int, user=Depends(_get_tecnico)):
    """Envia comando de limpeza ao dispenser-adapter. Bloqueado se slot está em operação."""
    if dispenser_id not in range(1, 7):
        raise HTTPException(400, "dispenser_id deve ser 1-6")

    with _lock:
        d_info   = _estado["dispensers"].get(str(dispenser_id), {})
        d_status = d_info.get("status", "idle")
        os_ativa = _estado["os_ativa"]

    # Bloqueio 1: OS ativa — nenhum dispenser pode ser limpo durante uma OS
    if os_ativa:
        os_atual = (os_ativa or {}).get("os_id", "?")
        raise HTTPException(
            status.HTTP_409_CONFLICT,
            f"Limpeza bloqueada: OS {os_atual} em andamento. "
            "Aguarde a conclusão da ordem de serviço.",
        )
    # Bloqueio 2: dispenser em operação ativa.
    # "concluido" NÃO entra: o slot já terminou a dispensa e pode ter residual
    # encalhado — limpar esse resto é justamente o propósito do botão da IHM.
    # (O simulador aplica a mesma regra em _do_limpar.)
    STATUS_BLOQUEADOS = {"carregando", "pronto", "dispensando", "aguardando_carga"}
    if d_status in STATUS_BLOQUEADOS:
        med = d_info.get("medicamento", f"Dispenser {dispenser_id}")
        raise HTTPException(
            status.HTTP_409_CONFLICT,
            f"Dispenser {dispenser_id} ({med}) em operação (status: '{d_status}'). "
            "Aguarde o sistema finalizar antes de limpar.",
        )

    ok = await orch.cmd_limpar(dispenser_id, user["sub"])
    if not ok:
        raise HTTPException(503, "Dispenser-adapter indisponível")

    await asyncio.to_thread(
        salvar_manutencao,
        "limpeza_dispenser",
        f"dispenser_{dispenser_id}",
        f"Limpeza manual por {user['sub']}",
        user["sub"],
    )
    _log("limpeza_solicitada", f"Limpeza D{dispenser_id} por {user['sub']}")
    return {"ok": True, "dispenser_id": dispenser_id,
            "msg": "Comando enviado. Aguardando confirmação do dispenser."}


@app.get("/manutencao/usuarios")
def listar_usuarios(user=Depends(_get_admin)):
    return get_usuarios()


@app.post("/manutencao/usuarios")
def criar_novo_usuario(req: UsuarioReq, user=Depends(_get_admin)):
    resultado = criar_usuario(req.username, req.senha, req.nome_completo, req.role)
    if not resultado.get("ok"):
        raise HTTPException(409, resultado.get("erro", "Erro ao criar usuário"))
    return resultado


@app.put("/manutencao/usuarios/{username}")
def editar_usuario(username: str, req: UsuarioUpdateReq, user=Depends(_get_admin)):
    resultado = atualizar_usuario(username, req.nome_completo, req.role, req.nova_senha)
    if not resultado.get("ok"):
        raise HTTPException(400, resultado.get("erro", "Erro ao atualizar"))
    # Role nova precisa valer AGORA, não daqui a AUTH_CACHE_TTL_S.
    _invalidar_cache_usuario(username)
    return resultado


@app.put("/manutencao/usuarios/{username}/desativar")
def desativar_usuario(username: str, user=Depends(_get_admin)):
    if username == user["sub"]:
        raise HTTPException(400, "Não pode desativar a própria conta")
    resultado = toggle_usuario_ativo(username, False)
    # Desativação é a operação em que o atraso do cache mais custa: é o botão
    # que o supervisor aperta quando quer alguém FORA agora.
    _invalidar_cache_usuario(username)
    return resultado


@app.put("/manutencao/usuarios/{username}/ativar")
def ativar_usuario(username: str, user=Depends(_get_admin)):
    resultado = toggle_usuario_ativo(username, True)
    _invalidar_cache_usuario(username)
    return resultado


# ── WebSocket ──────────────────────────────────────────────────────────────────
@app.websocket("/ws")
async def ws_endpoint(ws: WebSocket):
    await _ws_manager.connect(ws)
    with _lock:
        # deepcopy pelo mesmo motivo de `get_estado`: o json.dumps abaixo roda
        # fora do lock e percorreria os dicionários vivos.
        snap = copy.deepcopy(_estado)
    await ws.send_text(json.dumps({"tipo": "estado", **snap}, default=str))
    try:
        while True:
            await ws.receive_text()
    except WebSocketDisconnect:
        _ws_manager.disconnect(ws)
